from __future__ import annotations

import re
from typing import List, Tuple, Optional, Dict, Any, Callable

from openpyxl.workbook.workbook import Workbook

from llm_clients import EITranslationClient


# Any Devanagari character (Hindi script)
DEVANAGARI_CHAR_RE = re.compile(r"[\u0900-\u097F]")

# Protected tokens we MUST preserve exactly (and never translate)
# - <br>, <br/>, <br />  (case-insensitive)
# - [blank_1], [blank_2], etc.
PROTECTED_RE = re.compile(r"(<br\s*/?>|\[blank_\d+\])", re.IGNORECASE)


def contains_hindi(text: str) -> bool:
    return bool(DEVANAGARI_CHAR_RE.search(text))


def split_by_protected_tokens(text: str) -> List[Tuple[str, bool]]:
    """
    Split text into segments where protected tokens are isolated.
    Returns list of (segment, is_protected).
    """
    parts: List[Tuple[str, bool]] = []
    last = 0
    for m in PROTECTED_RE.finditer(text):
        if m.start() > last:
            parts.append((text[last:m.start()], False))
        parts.append((m.group(0), True))
        last = m.end()
    if last < len(text):
        parts.append((text[last:], False))
    return parts


def normalize_class(value: Any) -> str:
    if value is None:
        return ""
    s = str(value).strip().lower()
    s = s.replace("class", "").replace("grade", "").replace("std", "").replace("standard", "")
    return s.strip()


def build_row_prompt(base_prompt: str, class_value: str, use_class_guidance: bool) -> str:
    base_prompt = (base_prompt or "").strip()

    if not use_class_guidance:
        return base_prompt

    if class_value:
        return base_prompt + f"\n\nAudience: Students in Class {class_value}. Use simple age-appropriate English."

    return base_prompt + "\n\nAudience: School students. Use simple English."


def translate_cell_preserving_tokens(
    text: str,
    *,
    prompt: str,
    dictionary_override: Dict[str, str],
    client: Optional[EITranslationClient],
    cache: Dict[str, str],
) -> str:
    """
    Key fix:
    - Do NOT split by Hindi letters.
    - Split only by protected tokens (<br>, [blank_n]).
    - Translate whole spans that contain Hindi, so the model sees the full sentence/context.
    """
    if not contains_hindi(text):
        return text

    segments = split_by_protected_tokens(text)
    out_parts: List[str] = []

    for seg, is_protected in segments:
        if not seg:
            out_parts.append(seg)
            continue

        if is_protected:
            # Keep tokens exactly
            out_parts.append(seg)
            continue

        # Non-protected span
        if not contains_hindi(seg):
            out_parts.append(seg)
            continue

        # Dictionary override ONLY if the span itself is exactly the key (common for blank-drag words)
        trimmed = seg.strip()
        if trimmed in dictionary_override and trimmed == seg:
            out_parts.append(dictionary_override[trimmed])
            continue

        # Cache by exact span (including punctuation/spaces)
        if seg in cache:
            out_parts.append(cache[seg])
            if client:
                client.cache_hits += 1
            continue

        if client is None:
            cache[seg] = seg
            out_parts.append(seg)
            continue

        try:
            client.ai_calls += 1
            translated = client.translate_hi_to_en(seg, prompt=prompt)
            translated = translated.strip() or seg
        except Exception:
            translated = seg

        cache[seg] = translated
        out_parts.append(translated)

    return "".join(out_parts)


def _should_skip_column(header: str) -> bool:
    """
    Skip columns that are very unlikely to contain Hindi to translate.
    (Header-based rules only — no sampling.)
    """
    h = (header or "").strip().lower()
    if not h:
        return False

    exact_skips = {
        "subject", "context", "skill", "diff level", "template (type/drag)",
        "template", "class"
    }
    if h in exact_skips:
        return True

    suffix_skips = (
        "_allowed_char_set",
        "_min_chars_allowed",
        "_max_chars_allowed",
        "_size",
    )
    if h.endswith(suffix_skips):
        return True

    return False


def _find_effective_last_row(ws) -> int:
    """
    Data is contiguous: find the first row (starting from row 2) that is fully empty.
    Return the row just before it. If no empty row found, return ws.max_row.

    This prevents scanning "ghost rows" where Excel thinks formatting exists.
    """
    max_row = ws.max_row
    max_col = ws.max_column

    for r in range(2, max_row + 1):
        row_has_any = False
        for c in range(1, max_col + 1):
            v = ws.cell(row=r, column=c).value
            if v is None:
                continue
            if isinstance(v, str) and not v.strip():
                continue
            row_has_any = True
            break

        if not row_has_any:
            return r - 1

    return max_row


def _discover_candidate_columns(
    ws,
    *,
    last_row: int,
    progress_cb: Optional[Callable[[Dict[str, Any]], None]] = None,
) -> List[int]:
    """
    Pass 1: Identify columns that contain ANY Hindi anywhere (rows 2..last_row).
    Stops scanning a column once Hindi is found. No sampling bias.
    """
    max_col = ws.max_column

    col_names: Dict[int, str] = {}
    for c in range(1, max_col + 1):
        v = ws.cell(row=1, column=c).value
        col_names[c] = str(v).strip() if isinstance(v, str) and v.strip() else f"Col {c}"

    candidate_cols: List[int] = []
    scanned_cells = 0
    total_cells_est = max_col * max(0, (last_row - 1))

    def emit(event: Dict[str, Any]) -> None:
        if not progress_cb:
            return
        event.setdefault("event", "discover")
        event.setdefault("cells_done", scanned_cells)
        event.setdefault("cells_total", total_cells_est)
        progress_cb(event)

    emit({"phase": "discover_start", "effective_last_row": last_row})

    for c in range(1, max_col + 1):
        header = col_names.get(c, f"Col {c}")
        if _should_skip_column(header):
            continue

        found = False
        for r in range(2, last_row + 1):
            scanned_cells += 1
            v = ws.cell(row=r, column=c).value
            if isinstance(v, str) and v and (not v.startswith("=")) and contains_hindi(v):
                found = True
                break

            if progress_cb and (scanned_cells % 5000 == 0):
                emit({"phase": "discover_progress"})

        if found:
            candidate_cols.append(c)
            emit({"phase": "discover_found", "col": c, "col_name": header})

    emit({
        "phase": "discover_done",
        "candidate_cols": candidate_cols,
        "candidate_col_names": [col_names[c] for c in candidate_cols],
        "effective_last_row": last_row,
    })
    return candidate_cols


def process_workbook_inplace(
    wb: Workbook,
    *,
    base_prompt: str,
    dictionary_override: Dict[str, str],
    client: Optional[EITranslationClient],
    use_class_guidance: bool = True,
    sheet_name: Optional[str] = None,
    progress_cb: Optional[Callable[[Dict[str, Any]], None]] = None,
    update_every_cells: int = 300,
) -> None:
    """
    Two-pass strategy:
      1) Find effective_last_row (first fully empty row - 1)
      2) Discover candidate columns (no sampling bias)
      3) Translate only candidate columns up to effective_last_row

    IMPORTANT: Translation is span-level, preserving <br> and [blank_n] tokens.
    """
    ws = wb[sheet_name] if sheet_name else wb.active

    # column header names
    col_names: Dict[int, str] = {}
    for c in range(1, ws.max_column + 1):
        v = ws.cell(row=1, column=c).value
        col_names[c] = str(v).strip() if isinstance(v, str) and v.strip() else f"Col {c}"

    # class column
    class_col_idx = None
    for c in range(1, ws.max_column + 1):
        if ws.cell(row=1, column=c).value == "Class":
            class_col_idx = c
            break

    effective_last_row = _find_effective_last_row(ws)

    # nothing to do if sheet has no data rows
    if effective_last_row < 2:
        if progress_cb:
            progress_cb({"event": "candidates", "candidate_columns": []})
        return

    # discover candidate columns
    candidate_cols = _discover_candidate_columns(ws, last_row=effective_last_row, progress_cb=progress_cb)

    if progress_cb:
        progress_cb({
            "event": "candidates",
            "candidate_columns": [col_names[c] for c in candidate_cols],
            "effective_last_row": effective_last_row,
        })

    if not candidate_cols:
        return

    total_cells = max(0, (effective_last_row - 1) * len(candidate_cols))
    done_cells = 0
    cache: Dict[str, str] = {}

    def emit(event: Dict[str, Any]) -> None:
        if not progress_cb:
            return
        event.setdefault("cells_done", done_cells)
        event.setdefault("cells_total", total_cells)
        if client is not None:
            event.setdefault("ai_calls", client.ai_calls)
            event.setdefault("cache_hits", client.cache_hits)
        else:
            event.setdefault("ai_calls", 0)
            event.setdefault("cache_hits", 0)
        progress_cb(event)

    emit({"event": "progress"})

    try:
        for r in range(2, effective_last_row + 1):
            class_value = ""
            if class_col_idx:
                class_value = normalize_class(ws.cell(row=r, column=class_col_idx).value)
            row_prompt = build_row_prompt(base_prompt, class_value, use_class_guidance)

            for c in candidate_cols:
                done_cells += 1
                if update_every_cells > 0 and (done_cells % update_every_cells == 0):
                    emit({"event": "progress"})

                cell = ws.cell(row=r, column=c)
                v = cell.value
                if v is None or not isinstance(v, str) or not v or v.startswith("="):
                    continue

                if not contains_hindi(v):
                    continue

                new_v = translate_cell_preserving_tokens(
                    v,
                    prompt=row_prompt,
                    dictionary_override=dictionary_override,
                    client=client,
                    cache=cache,
                )

                if new_v != v:
                    cell.value = new_v
                    emit({
                        "event": "changed",
                        "row": r,
                        "col": c,
                        "col_name": col_names.get(c, f"Col {c}"),
                        "before": v,
                        "after": new_v,
                    })

        emit({"event": "progress"})

    except Exception as e:
        emit({
            "event": "error",
            "error": f"{type(e).__name__}: {e}",
            "http_debug": getattr(client, "last_http_debug", None) if client else None,
        })
        raise
